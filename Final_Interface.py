from tkinter import *
from tkinter import ttk
import openpyxl
from PIL import ImageTk, Image
import random 
import cv2
import customtkinter
import xlsxwriter
from datetime import date
from tkinter import messagebox

class StarterInterface:
    def __init__(self, interface):
        self.interface = interface
        interface.title("Graphical User Interface Starter")
        
        customtkinter.set_appearance_mode("light")  # Modes: system (default), light, dark
        customtkinter.set_default_color_theme("dark-blue")  # Themes: blue (default), dark-blue, green

        screen_width = self.interface.winfo_screenwidth()
        screen_height = self.interface.winfo_screenheight()

        self.user_id = random.randint(1000, 2000)
        window_width = int(screen_width)
        window_height = int(screen_height)

        x = int((screen_width - window_width) / 2)
        y = int((screen_height - window_height) / 2)
        self.interface.geometry('{}x{}+{}+{}'.format(window_width, window_height, x, y))

        self.fade_in()

    def fade_in(self):
        alpha = 0
        def fade_in_loop():
            nonlocal alpha
            alpha += 0.07
            self.interface.attributes("-alpha", alpha)
            self.interface.update()
            if alpha < 1:
                self.interface.after(10, fade_in_loop)

        fade_in_loop()
        self.display()

    def fade_out(self):
        alpha = 1
        def fade_out_loop():
            nonlocal alpha
            alpha -= 0.07
            self.interface.attributes("-alpha", alpha)
            self.interface.update()
            if alpha > 0:
                self.interface.after(10, fade_out_loop)
            else:
                self.interface.destroy()

        fade_out_loop()

    def display(self):
        self.label = customtkinter.CTkLabel(self.interface, text="Quality of Experience (QoE) Subjective Test", font=("Times New Roman", 60))
        self.label.place(relx=0.5, rely=0 + 0.1, anchor=CENTER)

        self.img = Image.open("Bshield_rgb.png")
        resized_img = self.img.resize((self.img.width // 4, self.img.height // 4))
        self.tk_resized_img = ImageTk.PhotoImage(resized_img)
        self.label_img = Label(self.interface, image=self.tk_resized_img)
        self.label_img.place(relx=0.5, rely=0.5 - 0.12, anchor=CENTER)

        self.wb = openpyxl.load_workbook('Book1.xlsx')
        self.sheet = self.wb.active

        self.start = customtkinter.CTkLabel(self.interface, text="Please fill out the following.")
        self.start.place(relx=0.5 - 0.1, rely=0.5 + 0.09, anchor=CENTER)

        self.name_label = customtkinter.CTkLabel(self.interface, text="Full Name")
        self.name_label.place(relx=0.5 - 0.111, rely=0.5 + 0.12, anchor=CENTER)
        self.name_entry = customtkinter.CTkEntry(self.interface, width=300)
        self.name_entry.place(relx=0.5 - 0.01, rely=0.5 + 0.12, anchor=CENTER)

        
        self.age_label = customtkinter.CTkLabel(self.interface, text="Age")
        self.age_label.place(relx=0.5 - 0.1, rely=0.5 + 0.15, anchor=CENTER)
        self.age_entry = customtkinter.CTkEntry(self.interface, width=100)
        self.age_entry.place(relx=0.5 - 0.061, rely=0.5 + 0.15, anchor=CENTER)

        self.grade_label = customtkinter.CTkLabel(self.interface, text="Class")
        self.grade_label.place(relx=0.5 - 0.1, rely=0.5 + 0.18, anchor=CENTER)
        self.GradeCombobox = customtkinter.CTkComboBox(self.interface, values=self.get_options2(), width=200)
        self.GradeCombobox.bind("<<ComboboxSelected>>", self.comboclick)
        self.GradeCombobox.place(relx=0.5 - 0.035, rely=0.5 + 0.18, anchor=CENTER)

        self.race_label = customtkinter.CTkLabel(self.interface, text="Race")
        self.race_label.place(relx=0.5 - 0.1, rely=0.5 + 0.21, anchor=CENTER)
        self.RaceCombobox = customtkinter.CTkComboBox(self.interface, values=self.get_options1(), width=200)
        self.RaceCombobox.bind("<<ComboboxSelected>>", self.comboclick2)
        self.RaceCombobox.place(relx=0.5 - 0.035, rely=0.5 + 0.21, anchor=CENTER)

        self.submit_button = customtkinter.CTkButton(self.interface, width=200, text="Submit", command=self.submit)
        self.submit_button.place(relx=0.5 - 0.05, rely=0.5 + 0.24)

    def get_options1(self):
        options1 = [
            "White",
            "Black or African American",
            "Native American or Alaska Native",
            "Asian",
            "Native Hawaiian, Samoan, or other Pacific Island", 
            "Hispanic", 
            "Other",
            "I prefer not to answer"
        ]
        return options1
    
    def get_options2(self):
        options2 = [
            "Freshman",
            "Sophmore",
            "Junior", 
            "Senior"
        ]
        return options2
    
    def get_options3(self):
        options3 = [str(i) for i in range(16, 66)]
        
        return options3
    
    def comboclick(self, event):
        print("Selected:", self.GradeCombobox.get())
    
    def comboclick2(self, event):
        print("Selected" + self.RaceCombobox.get())
        
    def submit(self):
        name = self.name_entry.get()
        age = self.age_entry.get()
        
        try:
            age = int(age)
        except ValueError:
        # Show an error message and return without submitting
            messagebox.showerror("Invalid Age", "Please enter a number for your age.")
            return
        
        grade = self.GradeCombobox.get()
        race = self.RaceCombobox.get()
        row = (name, age, grade, race)
        self.sheet.append(row)
        self.wb.save('Book1.xlsx')
        self.fade_out()


class Main_Interface:
    def __init__(self, interface):
        self.interface = interface
        self.answers = []
        
        interface.title("Graphical User Interface Main")
        # Screen height and width 
        customtkinter.set_appearance_mode("light")  # Modes: system (default), light, dark
        customtkinter.set_default_color_theme("dark-blue")  # Themes: blue (default), dark-blue, green

        screen_width = self.interface.winfo_screenwidth()
        screen_height = self.interface.winfo_screenheight()

        # Set the size of the window to 80% of the screen width and height
        window_width = int(screen_width)
        window_height = int(screen_height)

        # Set the geometry of the window to center it on the screen
        x = int((screen_width - window_width) / 2)
        y = int((screen_height - window_height) / 2)
        self.interface.geometry('{}x{}+{}+{}'.format(window_width, window_height, x, y))

        self.label = customtkinter.CTkLabel(interface, text="Videos!!!", font=("Times New Roman", 60))
        self.label.pack()

        self.videos = ("Squirrel_badQ.mp4", "Video2.mp4", "Video3.mp4")

        self.canvas = Canvas(self.interface, width=window_width, height=window_height)
        self.canvas.place(relx=0.57, rely=0.57, anchor=CENTER)
        
        self.current_video = 0
        self.cap = cv2.VideoCapture(self.videos[self.current_video])

        
        today = date.today()  
        user_id = "UserID_" + str(random.randint(1000, 2000)) + "_" + str(today)
        wb = xlsxwriter.Workbook(user_id + ".xlsx")
        worksheet = wb.add_worksheet()

        book1_wb = openpyxl.load_workbook('Book1.xlsx')
        # Select the active sheet
        sheet = book1_wb.active
        # Find the last row
        last_row = sheet.max_row
        # Find the last column
        last_column = sheet.max_column

        sheet.cell(row=last_row, column=last_column, value=user_id)
        # Save the workbook
        book1_wb.save('Book1.xlsx')
        # Close the workbook
        book1_wb.close()


        row = 2
        worksheet.write('B1', 'Video Resolution')
        worksheet.write('C1', 'Frame Rate')
        worksheet.write('D1', 'Sharpness')
        worksheet.write('E1', 'Movement Fluidity')
        worksheet.write('F1', 'Quality Consistency')
        worksheet.write('H1', 'Overall')
        for video in self.videos:
            worksheet.write(row, 0, video)
            row += 1

        # Function to update video frames in the canvas
        def update_video():
            ret, frame = self.cap.read()
            if ret:
                # Convert the frame from OpenCV format to PIL format
                img = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                img = Image.fromarray(img)
                img = ImageTk.PhotoImage(img)

                # Update the canvas with the new frame
                self.canvas.delete("all")
                self.canvas.img = img
                self.canvas.create_image(0, 0, anchor=NW, image=img)
                self.img = img
            if self.cap.get(cv2.CAP_PROP_POS_FRAMES) == self.cap.get(cv2.CAP_PROP_FRAME_COUNT) - 1:
                self.submit_button.configure(state="normal")
            # Schedule the next update
            interface.after(33, update_video)

        # Function to play the next video
        def play_next_video():
            self.current_video += 1
            if self.current_video >= len(self.videos):
                self.current_video = 0
                self.interface.quit()
                
            if self.cap.isOpened():
                self.cap.release()
            self.cap = cv2.VideoCapture(self.videos[self.current_video])

            self.play_button.place_forget()

            if self.current_video == len(self.videos) - 1:
                self.submit_button.configure(state="normal")  # Enable submit button
            else:
                self.submit_button.configure(state="disabled")  # Disable submit button

            # Check if it's the last video
            if self.current_video == len(self.videos) - 1:
                self.answers.append({
                    'Video Resolution': self.video_Resolution.get(),
                    'Frame Rate': self.frame_Rate.get(),
                    'Sharpness': self.sharpness.get(),
                    'Movement Fluidity': self.movement_Fluidity.get(), 
                    'Quality Consistency': self.quality_Consistency.get(),
                    'Overall': self.overall.get()
            })

        # Create buttons for play, next video, and quit
        self.play_button = customtkinter.CTkButton(interface, text="Play", command=update_video)
        self.play_button.place(relx=0.2, rely=0.8, anchor=CENTER)
        

        def replay_video():
            if self.cap.isOpened():
                self.cap.release()
            self.cap = cv2.VideoCapture(self.videos[self.current_video])
            self.submit_button.configure(state="disabled")  # Disable submit button
            update_video()  # Start playing the video again
        
        

        self.replay_button = customtkinter.CTkButton(interface, text="Replay", command=replay_video)
        self.replay_button.place(relx=0.4, rely=0.8, anchor=CENTER)
        self.replay_button.configure(state="normal")  # Disable replay button initially

        if self.current_video == len(self.videos) - 1:
            self.submit_button.configure(state="normal")  # Enable submit button
        else:
            self.replay_button.configure(state="normal")

        self.x = customtkinter.CTkLabel(interface, text="Please fill out the following to the best of your ability.")
        self.x.place(relx=0.5 + 0.35, rely=0.1, anchor=CENTER)

        self.video_Resolution_label = customtkinter.CTkLabel(interface, text="Video Resolution")
        self.video_Resolution_label.place(relx=0.5 + 0.3, rely=0.15, anchor=CENTER)
        self.video_Resolution = customtkinter.CTkComboBox(interface, values=self.score_options(), width=200)
        self.video_Resolution.bind("<<ComboboxSelected>>", self.comboclick)
        self.video_Resolution.place(relx=0.5 + 0.4, rely=0.15, anchor=CENTER)
        self.video_Resolution.set("-")

        self.frame_Rate_label = customtkinter.CTkLabel(interface, text="Frame_Rate")
        self.frame_Rate_label.place(relx=0.5 + 0.3, rely=0.18, anchor=CENTER)
        self.frame_Rate = customtkinter.CTkComboBox(interface, values=self.score_options(), width=200)
        self.frame_Rate.bind("<<ComboboxSelected>>", self.comboclick)
        self.frame_Rate.place(relx=0.5 + 0.4, rely=0.18, anchor=CENTER)
        self.frame_Rate.set("-")

        self.sharpness_label = customtkinter.CTkLabel(interface, text="Sharpness")
        self.sharpness_label.place(relx=0.5 + 0.3, rely=0.21, anchor=CENTER)
        self.sharpness = customtkinter.CTkComboBox(interface, values=self.score_options(), width=200)
        self.sharpness.bind("<<ComboboxSelected>>", self.comboclick)
        self.sharpness.place(relx=0.5 + 0.4, rely=0.21, anchor=CENTER)
        self.sharpness.set("-")

        self.movement_Fluidity_label = customtkinter.CTkLabel(interface, text="Movement Fluidity")
        self.movement_Fluidity_label.place(relx=0.5 + 0.3, rely=0.24, anchor=CENTER)
        self.movement_Fluidity = customtkinter.CTkComboBox(interface, values=self.score_options(), width=200)
        self.movement_Fluidity.bind("<<ComboboxSelected>>", self.comboclick)
        self.movement_Fluidity.place(relx=0.5 + 0.4, rely=0.24, anchor=CENTER)
        self.movement_Fluidity.set("-")


        self.quality_Consistency_label = customtkinter.CTkLabel(interface, text="Quality Consistency")
        self.quality_Consistency_label.place(relx=0.5 + 0.3, rely=0.27, anchor=CENTER)
        self.quality_Consistency = customtkinter.CTkComboBox(interface, values=self.score_options(), width=200)
        self.quality_Consistency.bind("<<ComboboxSelected>>", self.comboclick)
        self.quality_Consistency.place(relx=0.5 + 0.4, rely=0.27, anchor=CENTER)
        self.quality_Consistency.set("-")

        self.overall_label = customtkinter.CTkLabel(interface, text="Overall")
        self.overall_label.place(relx=0.5 + 0.3, rely=0.35, anchor=CENTER)
        self.overall = customtkinter.CTkComboBox(interface, values=self.score_options(), width=200)
        self.overall.bind("<<ComboboxSelected>>", self.comboclick)
        self.overall.place(relx=0.5 + 0.4, rely=0.35, anchor=CENTER)
        self.overall.set("-")

        self.num_videos = len(self.videos)

       
        def submit_scores():
            # Get the selected score values
            video_Resolution = self.video_Resolution.get()
            frame_Rate = self.frame_Rate.get()
            sharpness = self.sharpness.get()
            movement_Fluidity = self.movement_Fluidity.get()
            quality_Consistency = self.quality_Consistency.get()
            overall = self.overall.get()

            # Write the scores to the Excel file
            row = self.current_video + 2
            worksheet.write(row, 1, video_Resolution)
            worksheet.write(row, 2, frame_Rate)
            worksheet.write(row, 3, sharpness)
            worksheet.write(row, 4, movement_Fluidity)
            worksheet.write(row, 5, quality_Consistency)
            worksheet.write(row, 7, overall)

            self.answers.append({
                'Video Resolution': video_Resolution,
                'Frame Rate': frame_Rate,
                'Sharpness': sharpness,
                'Movement Fluidity': movement_Fluidity, 
                'Quality Consistency': quality_Consistency,
                 'Overall': overall
            })

            if self.current_video == len(self.videos) - 1:
                # Close the workbook when all responses are collected
                wb.close()  # Close the workbook and save changes
                self.show_thank_you()

            play_next_video()
            self.video_Resolution.set('-')
            self.frame_Rate.set('-')
            self.sharpness.set('-')
            self.movement_Fluidity.set('-')
            self.quality_Consistency.set('-')
            self.overall.set('-')

            if self.current_video == len(self.videos) - 1:
                self.submit_button.configure(state="disabled")
            
        

        self.submit_button = customtkinter.CTkButton(interface, text="Submit Scores", command=submit_scores)
        self.submit_button.place(relx=0.5 + 0.4, rely=0.45, anchor=CENTER)
        self.submit_button.configure(state="disabled")

    def comboclick(self, event):
        print("Selected:", self.videoQuality.get())

    def score_options(self):
        scoreOptions = [
            "1",
            "2",
            "3",
            "4",
            "5",
            "6",
            "7",
            "8",
            "9"
        ]
        return scoreOptions

    def show_thank_you(self):
        # Create a new window for the thank you screen
        messagebox.showinfo("Thank You", "Thank you for your submissions!")




def starter_main():
    root = Tk()
    g = StarterInterface(root)
    root.mainloop()

def main():
    root = Tk()
    x = Main_Interface(root)
    root.mainloop()

if __name__ == '__main__':
    starter_main()
    main() 

